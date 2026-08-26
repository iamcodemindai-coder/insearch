import json
import os
import re
import time
from pathlib import Path

from openpyxl import Workbook
from dotenv import load_dotenv
from openai import OpenAI


# ============================================================
# CONFIG
# ============================================================

load_dotenv()

API_KEY = os.getenv("OPENAI_API_KEY")
BASE_URL = os.getenv("OPENAI_BASE_URL")
MODEL = os.getenv("OPENAI_MODEL", "gpt-5")

if not API_KEY:
    raise RuntimeError("OPENAI_API_KEY is missing in .env")


client_config = {"api_key": API_KEY}

if BASE_URL:
    client_config["base_url"] = BASE_URL

client = OpenAI(**client_config)


# ============================================================
# SYSTEM PROMPT
# ============================================================

SYSTEM_PROMPT = """
You are a high-precision semantic similarity evaluator.

Your task is to compare ONE CURRENT ISSUE SUMMARY against
MULTIPLE HISTORICAL NOTIFICATIONS.

Return one score for EVERY notification.

The score represents semantic relevance between the complete
meaning of the current issue and the complete meaning of the
historical notification.

============================================================
IMPORTANT: GENERIC VS SPECIFIC INFORMATION
============================================================

A notification can contain:

1. Generic information
2. Specific technical/problem information

Example:

"damaged material (damaged lead screw)"

Generic:
"damaged material"

Specific:
"damaged lead screw"

The generic part should NOT dominate the score.

However, do NOT completely ignore the generic part either.

Use the COMPLETE TEXT.

============================================================
CRITICAL GENERIC CASE
============================================================

Current issue:

"damaged material"

Historical notifications:

"damaged material (misaligned hook talons)"
"damaged material (soldering issues)"
"damaged material (high SMA wire resistance)"
"damaged material (damaged lead screw)"

The current issue is generic-only.

Therefore:

- Do NOT assume that any specific historical subtype is the
  actual current problem.
- Do NOT give these records 90+ merely because they contain
  "damaged material".
- BUT also do NOT force every record to exactly the same score.
- Evaluate the complete semantic content of each notification.
- Preserve meaningful differences in semantic closeness.
- Generic category similarity contributes to relevance.
- Specific information refines the score.

The goal is RELATIVE semantic relevance, not keyword counting.

============================================================
VERY IMPORTANT: DO NOT FLATTEN SCORES
============================================================

Do NOT return:

25, 25, 25, 25

just because all notifications contain the same generic phrase.

If the complete meanings have different semantic closeness,
their scores should be different.

For example, a reasonable result may look like:

65
61
68
73

rather than:

25
25
25
25

Exact values are your judgement.

============================================================
SPECIFIC MATCH
============================================================

If both records contain the same specific problem:

Current:
"damaged material (soldering issue)"

Historical:
"damaged material (soldering problem)"

=> HIGH score.

If specific problems are different:

Current:
"damaged material (soldering issue)"

Historical:
"damaged material (battery corrosion)"

=> LOW/MODERATE score.

The shared generic phrase alone is not enough for a high score.

============================================================
SEMANTIC UNDERSTANDING
============================================================

Understand:

- synonyms
- paraphrases
- technical terminology
- abbreviations
- equivalent descriptions
- word order
- grammar
- verb tense / grammatical mistakes
- complete technical meaning

GRAMMAR EQUIVALENCE:

Treat grammatically incorrect but semantically equivalent
technical phrases as the same meaning.

Example:

"Autoinjector did not activate"
and
"Autoinjector did not activated"

describe the same technical event.

Therefore, a grammar/verb-form difference alone should NOT
significantly reduce the score.

============================================================
CRITICAL: NEGATION AND POLARITY
============================================================

Negation, success/failure, enabled/disabled, present/absent,
working/not working, activated/not activated and similar
opposite states are CRITICAL semantic information.

Do NOT calculate similarity mainly from shared nouns and verbs.

Example:

Current:
"Autoinjector activated"

Historical:
"Autoinjector did not activate"

These describe OPPOSITE outcomes.

"activated" = successful activation
"did not activate" = failed/non-activation

Therefore, this pair MUST receive a LOW similarity score,
even though "Autoinjector" and "activate" are shared.

Do NOT give a high score merely because the component and action
words are the same when the outcome/polarity is opposite.

Examples of equivalent NEGATIVE meaning:

"Autoinjector did not activate"
"Autoinjector did not activated"
"Autoinjector failed to activate"
"Autoinjector was unable to activate"

These describe the same negative event and should receive a
HIGH similarity score.

Examples of equivalent POSITIVE meaning:

"Autoinjector activated"
"Autoinjector successfully activated"
"Autoinjector was activated"

These describe the same positive event and should receive a
HIGH similarity score.

If one text describes success and the other describes failure,
the score MUST be low regardless of word overlap.

Always determine the EVENT OUTCOME / POLARITY before assigning
a high score.

============================================================
SCORING GUIDELINES
============================================================

0.95-1.00:
Almost identical complete meaning.

0.85-0.94:
Very strong semantic match.

0.70-0.84:
Strong semantic match.

0.55-0.69:
Moderate-to-strong semantic relationship.

0.40-0.54:
Moderate relationship.

0.25-0.39:
Weak relationship / mostly generic relationship.

0.10-0.24:
Very weak relationship.

0.00-0.09:
Essentially unrelated.

IMPORTANT POLARITY OVERRIDE:

If the only major difference is opposite outcome/polarity
(for example "activated" vs "did not activate"), do NOT use
the high lexical overlap to produce a high score. Such pairs
must remain in the low similarity range unless the surrounding
context clearly makes the events equivalent.

IMPORTANT:

These are guidelines, NOT fixed buckets.

Do not automatically assign the same score to records in
the same category.

Do not automatically assign 25%.

============================================================
EMBEDDING-LIKE RELATIVE BEHAVIOUR
============================================================

The historical records may all share a generic phrase.

Still compare their complete text and preserve relative
semantic closeness.

If one notification is semantically closer to the current
issue than another, its score should normally be higher.

The purpose is to produce a meaningful relevance ranking.

Do NOT try to reproduce a specific embedding cosine value.

For example:

Embedding scores:

0.64
0.64
0.67
0.73

The LLM does NOT need to output exactly:

64
64
67
73

But it should preserve meaningful relative differences when
the text supports them.

============================================================
MISSING INFORMATION
============================================================

Never invent information.

If the current issue does not mention a specific component,
do not assume that component is the current problem.

============================================================
BLANK NOTIFICATIONS
============================================================

If pcm_inv_notif is:

null
""
whitespace
or missing completely

the score MUST be exactly 0.

Do not fail.

============================================================
OUTPUT
============================================================

Return ONLY valid JSON:

{
    "results": [
        {
            "pcm_inv_notif_id": "123",
            "score": 0.92
        }
    ]
}

Rules:

- one result for every notification
- preserve exact notification ID
- score must be a numeric value between 0.0 and 1.0
- blank/missing notification = 0
- no explanation
- no markdown
- no additional fields
"""


# ============================================================
# HELPERS
# ============================================================

def normalize(text):

    if text is None:
        return ""

    return re.sub(
        r"\s+",
        " ",
        str(text).strip()
    )


def load_json(path):

    try:

        with open(
            path,
            "r",
            encoding="utf-8"
        ) as f:

            return json.load(f)

    except FileNotFoundError:

        raise RuntimeError(
            f"Input file not found: {path}"
        )

    except json.JSONDecodeError as e:

        raise RuntimeError(
            f"Invalid JSON in {path.name}: {e}"
        )


def save_json(data, path):

    with open(
        path,
        "w",
        encoding="utf-8"
    ) as f:

        json.dump(
            data,
            f,
            indent=4,
            ensure_ascii=False
        )


# ============================================================
# EXTRACT NOTIFICATIONS
# ============================================================

def get_notifications(search_results):

    notifications = []

    for index, item in enumerate(search_results):

        if not isinstance(item, dict):

            print(
                f"COMMENT: search_results[{index}] "
                f"is invalid. Skipping."
            )

            continue

        notif_id = item.get(
            "pcm_inv_notif_id"
        )

        if notif_id is not None:

            notif_id = normalize(
                notif_id
            )

        text = normalize(
            item.get("pcm_inv_notif")
        )

        notifications.append({
            "index": index,
            "id": notif_id,
            "text": text,
            "missing": not bool(text)
        })

        if not text:

            print(
                f"COMMENT: Notification missing/blank "
                f"for ID {notif_id}. Score = 0."
            )

    return notifications


# ============================================================
# LLM SIMILARITY
# ============================================================

def calculate_similarity(
    issue_summary,
    notifications
):

    valid_notifications = [
        x for x in notifications
        if x["text"] and x["id"]
    ]

    # --------------------------------------------------------
    # If nothing valid exists
    # --------------------------------------------------------

    if not valid_notifications:

        return {
            "results": [
                {
                    "pcm_inv_notif_id": x["id"],
                    "score": 0
                }
                for x in notifications
                if x["id"]
            ]
        }, 0.0

    # --------------------------------------------------------
    # Prepare records
    # --------------------------------------------------------

    records = []

    for x in valid_notifications:

        records.append({
            "pcm_inv_notif_id": x["id"],
            "pcm_inv_notif": x["text"]
        })

    user_prompt = f"""
CURRENT ISSUE SUMMARY:

{issue_summary}


HISTORICAL NOTIFICATIONS:

{json.dumps(
    records,
    indent=2,
    ensure_ascii=False
)}


TASK:

Evaluate EVERY historical notification independently against
the current issue.

Then calibrate the scores across the complete set so that
meaningful differences between notifications are preserved.

IMPORTANT:

If the current issue is:

"damaged material"

and notifications contain:

"damaged material (misaligned hook talons)"
"damaged material (soldering issues)"
"damaged material (high SMA wire resistance)"
"damaged material (damaged lead screw)"

do NOT return identical scores simply because
"damaged material" is repeated.

The generic category should contribute to similarity, while
the complete notification text should refine the score.

Do not assume any specific subtype is the actual current issue.

CRITICAL POLARITY CHECK:

Before scoring, determine whether the current issue and each
notification describe the same outcome.

For example:

Current: "Autoinjector activated"
Notification: "Autoinjector did not activate"

These are opposite outcomes. The shared words "Autoinjector"
and "activate" must NOT cause a high score. The score must be low.

But:

Current: "Autoinjector did not activate"
Notification: "Autoinjector did not activated"

These are the same negative event. The grammatical difference
must NOT reduce the score significantly.

Always prioritize complete event meaning and outcome over
simple word overlap.

Return one score for every notification.

Return ONLY JSON.
"""

    start = time.perf_counter()

    try:

        response = client.chat.completions.create(
            model=MODEL,
            messages=[
                {
                    "role": "system",
                    "content": SYSTEM_PROMPT
                },
                {
                    "role": "user",
                    "content": user_prompt
                }
            ],
            response_format={
                "type": "json_object"
            }
        )

    except Exception as e:

        raise RuntimeError(
            f"LLM API call failed: {e}"
        )

    llm_time = time.perf_counter() - start

    # --------------------------------------------------------
    # Read response
    # --------------------------------------------------------

    content = (
        response
        .choices[0]
        .message
        .content
    )

    if not content:

        raise RuntimeError(
            "LLM returned empty response."
        )

    content = content.strip()

    content = re.sub(
        r"^```(?:json)?",
        "",
        content,
        flags=re.IGNORECASE
    )

    content = re.sub(
        r"```$",
        "",
        content
    ).strip()

    try:

        result = json.loads(content)

    except json.JSONDecodeError as e:

        raise RuntimeError(
            f"Invalid JSON returned by LLM:\n{content}"
        ) from e

    if not isinstance(result, dict):

        raise RuntimeError(
            "LLM response must be an object."
        )

    results = result.get("results")

    if not isinstance(results, list):

        raise RuntimeError(
            "LLM response must contain 'results' list."
        )

    # --------------------------------------------------------
    # Validate IDs and scores
    # --------------------------------------------------------

    expected_ids = {
        x["id"]
        for x in valid_notifications
    }

    received = {}

    for item in results:

        if not isinstance(item, dict):

            raise RuntimeError(
                "Invalid result object returned by LLM."
            )

        notif_id = normalize(
            item.get("pcm_inv_notif_id")
        )

        if not notif_id:

            raise RuntimeError(
                "LLM returned result without notification ID."
            )

        if notif_id not in expected_ids:

            raise RuntimeError(
                f"Unexpected notification ID: {notif_id}"
            )

        if notif_id in received:

            raise RuntimeError(
                f"Duplicate notification ID: {notif_id}"
            )

        try:

            score = float(
                item.get("score")
            )

        except (
            TypeError,
            ValueError
        ):

            raise RuntimeError(
                f"Invalid score for {notif_id}"
            )

        if not 0.0 <= score <= 1.0:

            raise RuntimeError(
                f"Score out of range for {notif_id}: {score}"
            )

        received[notif_id] = round(
            score,
            4
        )

    # --------------------------------------------------------
    # Make sure every valid notification got score
    # --------------------------------------------------------

    missing = (
        expected_ids -
        set(received)
    )

    if missing:

        raise RuntimeError(
            f"LLM missed notification IDs: {missing}"
        )

    # --------------------------------------------------------
    # Rebuild results in ORIGINAL order
    # --------------------------------------------------------

    final_results = []

    for x in notifications:

        notif_id = x["id"]

        if not notif_id:
            continue

        if x["missing"]:

            score = 0

        else:

            score = received.get(
                notif_id,
                0
            )

        final_results.append({
            "pcm_inv_notif_id": notif_id,
            "score": score
        })

    return {
        "results": final_results
    }, llm_time


# ============================================================
# ADD SCORE TO ORIGINAL JSON
# ============================================================

def add_scores(
    original_data,
    similarity_result
):

    score_map = {
        str(x["pcm_inv_notif_id"]): float(x["score"])
        for x in similarity_result["results"]
        if x.get("pcm_inv_notif_id") is not None
    }

    output = dict(original_data)

    updated = []

    for item in original_data.get(
        "search_results",
        []
    ):

        if not isinstance(item, dict):

            updated.append(item)

            continue

        item_copy = {}

        notif_id = normalize(
            item.get("pcm_inv_notif_id")
        )

        score = score_map.get(
            notif_id,
            0
        )

        inserted = False

        for key, value in item.items():

            item_copy[key] = value

            if (
                key == "pcm_inv_notif"
                and not inserted
            ):

                item_copy["score"] = score

                inserted = True

        if not inserted:

            item_copy["score"] = score

        updated.append(item_copy)

    output["search_results"] = updated

    return output


# ============================================================
# MAIN PROCESS
# ============================================================

def process(data):

    if not isinstance(data, dict):

        raise ValueError(
            "Input JSON must be an object."
        )

    issue_summary = normalize(
        data.get("pcm_issue_summary")
    )

    if not issue_summary:

        raise ValueError(
            "pcm_issue_summary is missing/empty."
        )

    search_results = data.get(
        "search_results"
    )

    if not isinstance(
        search_results,
        list
    ):

        raise ValueError(
            "search_results must be a list."
        )

    notifications = get_notifications(
        search_results
    )

    print("\n" + "=" * 70)
    print("INPUT")
    print("=" * 70)

    print(
        f"Issue Summary : {issue_summary}"
    )

    print(
        f"Records       : {len(notifications)}"
    )

    # --------------------------------------------------------
    # LLM
    # --------------------------------------------------------

    similarity_result, llm_time = (
        calculate_similarity(
            issue_summary,
            notifications
        )
    )

    # --------------------------------------------------------
    # Add scores
    # --------------------------------------------------------

    output = add_scores(
        data,
        similarity_result
    )

    return output, llm_time


# ============================================================
# BATCH PROCESSING
# ============================================================

def run_batch(
    input_dir="batch_inputs",
    excel_file="batch_results.xlsx",
    output_dir="batch_outputs"
):

    input_dir = Path(input_dir)
    output_dir = Path(output_dir)

    # --------------------------------------------------------
    # Create folders if they don't exist
    # --------------------------------------------------------

    input_dir.mkdir(
        parents=True,
        exist_ok=True
    )

    output_dir.mkdir(
        parents=True,
        exist_ok=True
    )

    # --------------------------------------------------------
    # Find ALL JSON files
    # --------------------------------------------------------

    json_files = sorted(
        input_dir.glob("*.json")
    )

    if not json_files:

        print(
            f"No JSON files found in: {input_dir}"
        )

        print(
            "\nPut your JSON files inside "
            f"'{input_dir}' folder."
        )

        return

    # --------------------------------------------------------
    # Create Excel
    # --------------------------------------------------------

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Similarity Results"

    headers = [
        "Input File",
        "Issue Summary",
        "Issue ID",
        "Automation ID",
        "Notification Position",
        "Notification ID",
        "Notification",
        "Score",
        "Score %",
        "LLM Response Time (sec)",
        "Total Time (sec)",
        "Status"
    ]

    sheet.append(headers)

    # --------------------------------------------------------
    # Batch timer
    # --------------------------------------------------------

    batch_start = time.perf_counter()

    successful_files = 0
    failed_files = 0

    # --------------------------------------------------------
    # Process every JSON
    # --------------------------------------------------------

    for file_index, input_file in enumerate(
        json_files,
        start=1
    ):

        print("\n" + "=" * 80)
        print(
            f"BATCH {file_index}/{len(json_files)}"
        )
        print(
            f"FILE: {input_file.name}"
        )
        print("=" * 80)

        case_start = time.perf_counter()

        try:

            # ------------------------------------------------
            # Load JSON
            # ------------------------------------------------

            data = load_json(
                input_file
            )

            # ------------------------------------------------
            # Process similarity
            # ------------------------------------------------

            output, llm_time = process(
                data
            )

            # ------------------------------------------------
            # Save individual output JSON
            # ------------------------------------------------

            output_file = (
                output_dir /
                f"{input_file.stem}_output.json"
            )

            save_json(
                output,
                output_file
            )

            total_case_time = (
                time.perf_counter()
                - case_start
            )

            # ------------------------------------------------
            # Metadata
            # ------------------------------------------------

            issue_summary = normalize(
                data.get(
                    "pcm_issue_summary"
                )
            )

            issue_id = normalize(
                data.get(
                    "pcm_issue_id"
                )
            )

            automation_id = normalize(
                data.get(
                    "automation_id"
                )
            )

            # ------------------------------------------------
            # Excel rows
            # ------------------------------------------------

            search_results = output.get(
                "search_results",
                []
            )

            for position, item in enumerate(
                search_results,
                start=1
            ):

                if not isinstance(
                    item,
                    dict
                ):

                    continue

                notif_id = normalize(
                    item.get(
                        "pcm_inv_notif_id"
                    )
                )

                notification_text = normalize(
                    item.get(
                        "pcm_inv_notif"
                    )
                )

                score = item.get(
                    "score",
                    0.0
                )

                try:

                    score = float(score)

                except (
                    TypeError,
                    ValueError
                ):

                    score = 0.0

                score_percentage = (
                    score * 100
                )

                sheet.append([
                    input_file.name,
                    issue_summary,
                    issue_id,
                    automation_id,
                    position,
                    notif_id,
                    notification_text,
                    round(score, 4),
                    f"{score_percentage:.2f}%",
                    round(llm_time, 3),
                    round(total_case_time, 3),
                    "SUCCESS"
                ])

            successful_files += 1

            print(
                f"Issue Summary     : {issue_summary}"
            )

            print(
                f"Notifications     : {len(search_results)}"
            )

            print(
                f"LLM response time : {llm_time:.3f} sec"
            )

            print(
                f"Total time        : "
                f"{total_case_time:.3f} sec"
            )

            print(
                f"Output JSON       : {output_file}"
            )

        except Exception as e:

            failed_files += 1

            total_case_time = (
                time.perf_counter()
                - case_start
            )

            print(
                f"ERROR in {input_file.name}: {e}"
            )

            # ------------------------------------------------
            # Put failed file information in Excel
            # ------------------------------------------------

            sheet.append([
                input_file.name,
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                round(
                    total_case_time,
                    3
                ),
                f"ERROR: {e}"
            ])

    # ========================================================
    # EXCEL FORMATTING
    # ========================================================

    sheet.freeze_panes = "A2"

    sheet.auto_filter.ref = (
        sheet.dimensions
    )

    widths = {
        "A": 28,
        "B": 40,
        "C": 28,
        "D": 28,
        "E": 22,
        "F": 30,
        "G": 60,
        "H": 12,
        "I": 12,
        "J": 25,
        "K": 20,
        "L": 25
    }

    for column, width in widths.items():

        sheet.column_dimensions[
            column
        ].width = width

    # --------------------------------------------------------
    # Number formatting
    # --------------------------------------------------------

    for row in sheet.iter_rows(
        min_row=2,
        min_col=8,
        max_col=8
    ):

        for cell in row:

            cell.number_format = "0.0000"

    for row in sheet.iter_rows(
        min_row=2,
        min_col=10,
        max_col=11
    ):

        for cell in row:

            cell.number_format = "0.000"

    # --------------------------------------------------------
    # Save Excel
    # --------------------------------------------------------

    workbook.save(
        excel_file
    )

    # ========================================================
    # FINAL SUMMARY
    # ========================================================

    total_batch_time = (
        time.perf_counter()
        - batch_start
    )

    print("\n" + "=" * 80)
    print("BATCH COMPLETE")
    print("=" * 80)

    print(
        f"Total JSON files : {len(json_files)}"
    )

    print(
        f"Successful       : {successful_files}"
    )

    print(
        f"Failed           : {failed_files}"
    )

    print(
        f"Excel saved      : {excel_file}"
    )

    print(
        f"Output folder    : {output_dir}"
    )

    print(
        f"Total batch time : "
        f"{total_batch_time:.3f} sec"
    )

    print("=" * 80)


# ============================================================
# SINGLE FILE MODE
# ============================================================

def run_single():

    input_file = Path(
        "example_input.json"
    )

    output_file = Path(
        "output.json"
    )

    print("=" * 70)
    print("GENAI SEMANTIC SIMILARITY")
    print("=" * 70)

    total_start = time.perf_counter()

    data = load_json(
        input_file
    )

    output, llm_time = process(
        data
    )

    save_json(
        output,
        output_file
    )

    total_time = (
        time.perf_counter()
        - total_start
    )

    print("\n" + "=" * 70)
    print("SCORES")
    print("=" * 70)

    for item in output.get(
        "search_results",
        []
    ):

        score = item.get(
            "score",
            0
        )

        print(
            f"{item.get('pcm_inv_notif_id')} "
            f"-> {score} "
            f"({float(score) * 100:.2f}%)"
        )

    print("\n" + "=" * 70)
    print("PERFORMANCE")
    print("=" * 70)

    print(
        f"LLM response time : "
        f"{llm_time:.3f} sec"
    )

    print(
        f"Total time        : "
        f"{total_time:.3f} sec"
    )

    print(
        f"Output saved      : "
        f"{output_file}"
    )

    print("\n" + "=" * 70)
    print("FINAL JSON")
    print("=" * 70)

    print(
        json.dumps(
            output,
            indent=4,
            ensure_ascii=False
        )
    )


# ============================================================
# MAIN
# ============================================================

def main():

    batch_dir = Path(
        "batch_inputs"
    )

    # --------------------------------------------------------
    # If batch_inputs contains JSON files
    # -> automatically run batch mode
    # --------------------------------------------------------

    if batch_dir.exists() and list(
        batch_dir.glob("*.json")
    ):

        run_batch(
            input_dir="batch_inputs",
            excel_file="batch_results.xlsx",
            output_dir="batch_outputs"
        )

        return

    # --------------------------------------------------------
    # Otherwise run single JSON
    # --------------------------------------------------------

    run_single()


# ============================================================
# ENTRY POINT
# ============================================================

if __name__ == "__main__":
    main()